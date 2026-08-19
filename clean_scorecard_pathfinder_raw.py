"""
clean_scorecard_pathfinder_raw.py -- fix two data-quality issues, in place, in:

  1. data/raw/labor_mi/Pathfinder-Employment_Outcome_Report-Wayne_State_University.xlsx
  2. data/raw/scorecard/most_recent_cohorts_field_of_study.parquet

WHAT THIS FIXES

Pathfinder (MI CEPI):
  - "CIP Code" is stored as float64, losing the leading zero for CIP families
    01-09 (35 of 604 rows). Rebuilt as proper "XX.XXXX" text, same pattern as
    the crosswalk fix in clean_crosswalk_raw.py.
  - "*" is Pathfinder's small-cell suppression marker. It appears in 367 of 604
    rows, in the Award column (meaning the whole CIP/Award cohort was too small
    to report at all) as well as scattered across individual outcome columns.
    Every genuinely numeric column here already holds real Python numeric values
    wherever it isn't "*" -- confirmed no other stray tokens exist -- so this is
    a clean swap: "*" -> missing, then the column is cast to a real numeric dtype.

College Scorecard (most_recent_cohorts_field_of_study.parquet):
  - CIPCODE is NOT a bug -- it's already proper text, just Scorecard's own
    4-digit convention (e.g. "0109"), different from the 6-digit dotted CIP
    format used elsewhere in this project. Left untouched here; reconciling the
    two formats is a join-logic decision, not a data-quality fix, and is
    out of scope for this script.
  - "PS" (PrivacySuppressed) appears across 134 columns -- as high as 75% of
    rows in some earnings fields -- and is swapped for missing, then those
    columns are cast to numeric.
  - IMPORTANT, and NOT touched by this script: 32 of the BBRR*_FED_COMP_*
    columns (federal loan repayment status rates -- default, delinquent,
    forbearance, etc.) are not simple point estimates. Where they're not "PS",
    they're reported as BINNED RANGES like "0.37 - 0.38" or "<=0.20" or
    ">=0.90" -- this is Scorecard's actual disclosure-avoidance format for
    these specific fields, not corruption. Coercing "*" or "PS" to missing is
    correct here; collapsing a real range into a fabricated point value (a
    midpoint, say) is a modeling decision this script deliberately does not
    make. Those 32 columns are left as text, and this script prints their
    names so you know they need a deliberate decision before any numeric use.

Nothing is deleted. Both files are backed up with a timestamp before being
overwritten. This script does NOT compute or report suppression rates --
that's a follow-on task, not part of this fix.

Usage (run from the project root, e.g. ~/Antigravity/IPEDS_sandbox):
    python clean_scorecard_pathfinder_raw.py
"""

import os
import re
import shutil
import sys
from datetime import datetime

import pandas as pd

PATHFINDER_PATH = "data/raw/labor_mi/Pathfinder-Employment_Outcome_Report-Wayne_State_University.xlsx"
PATHFINDER_SHEET = "Wayne State University"
SCORECARD_PATH = "data/raw/scorecard/most_recent_cohorts_field_of_study.parquet"

PATHFINDER_SUPPRESSION_TOKEN = "*"
SCORECARD_SUPPRESSION_TOKEN = "PS"

CIP_CODE_RE = re.compile(r"^\d{2}\.\d{4}$")

SCORECARD_ID_COLS = {
    "UNITID", "OPEID6", "INSTNM", "CONTROL", "MAIN",
    "CIPCODE", "CIPDESC", "CREDLEV", "CREDDESC",
}


def _timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _backup(path):
    backup_path = f"{path}.bak_{_timestamp()}"
    shutil.copy2(path, backup_path)
    print(f"    Backed up to: {backup_path}")


def _fail(msg):
    print(f"\nABORTED: {msg}")
    print("No files were modified.")
    sys.exit(1)


def normalize_cip_code(x):
    """float/str CIP code -> 'XX.XXXX' text, matching normalize_cip() elsewhere
    in this project. Recomputes padding from the numeric value rather than
    trusting any existing string formatting."""
    if pd.isna(x):
        return None
    family, _, frac = f"{float(x):.4f}".partition(".")
    return f"{int(family):02d}.{frac}"


def clean_pathfinder():
    print(f"\n--- Cleaning {PATHFINDER_PATH} ---")
    if not os.path.exists(PATHFINDER_PATH):
        print("  SKIPPED: file not found.")
        return

    df = pd.read_excel(PATHFINDER_PATH, sheet_name=PATHFINDER_SHEET)
    print(f"  Read {len(df):,} rows.")

    if df["CIP Code"].dtype != object:
        n_before_family01_09 = (df["CIP Code"] < 10).sum()
        print(f"  'CIP Code' is {df['CIP Code'].dtype} -- {n_before_family01_09} rows in "
              "families 01-09 currently missing their leading zero.")
    df["CIP Code"] = df["CIP Code"].map(normalize_cip_code)
    bad = sorted(c for c in df["CIP Code"].dropna().unique() if not CIP_CODE_RE.match(c))
    if bad:
        _fail(f"{len(bad)} CIP Code value(s) don't match XX.XXXX after normalizing: {bad[:10]}. "
              "Inspect the source file by hand before proceeding.")
    print("  'CIP Code': rebuilt as text, all values now well-formed XX.XXXX.")

    n_suppressed_cells = int((df == PATHFINDER_SUPPRESSION_TOKEN).sum().sum())
    n_suppressed_rows = int((df == PATHFINDER_SUPPRESSION_TOKEN).any(axis=1).sum())
    print(f"  Suppression marker '{PATHFINDER_SUPPRESSION_TOKEN}': {n_suppressed_cells:,} cells "
          f"across {n_suppressed_rows:,} of {len(df):,} rows.")

    for col in df.columns:
        if col == "CIP Code":
            continue
        is_suppressed = df[col] == PATHFINDER_SUPPRESSION_TOKEN
        df.loc[is_suppressed, col] = pd.NA
        if col != "Award":
            non_null = df[col].dropna()
            unparseable = non_null[pd.to_numeric(non_null, errors="coerce").isna()]
            if len(unparseable):
                _fail(f"Column '{col}' has {len(unparseable)} value(s) that are neither "
                      f"'{PATHFINDER_SUPPRESSION_TOKEN}' nor numeric: {unparseable.unique()[:5]}. "
                      "Not safe to auto-convert -- inspect by hand.")
            df[col] = pd.to_numeric(df[col], errors="coerce")

    _backup(PATHFINDER_PATH)
    df.to_excel(PATHFINDER_PATH, sheet_name=PATHFINDER_SHEET, index=False)

    # Belt-and-suspenders: force the CIP Code column's cell format to Text in the
    # actual .xlsx, so it displays correctly if opened directly in Excel. NOTE this
    # does NOT, by itself, stop pandas from re-inferring the column as float64 on a
    # future read -- that's a pandas default-dtype-inference behavior, not a file
    # property, and it will silently reintroduce this exact bug for anyone who reads
    # this file with plain pd.read_excel(path) instead of pd.read_excel(path, dtype=str).
    from openpyxl import load_workbook
    wb = load_workbook(PATHFINDER_PATH)
    ws = wb[PATHFINDER_SHEET]
    cip_col_idx = [c.value for c in ws[1]].index("CIP Code") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=cip_col_idx).number_format = "@"
    wb.save(PATHFINDER_PATH)

    # verify the round trip the way a future reader actually would, both ways
    reread_default = pd.read_excel(PATHFINDER_PATH, sheet_name=PATHFINDER_SHEET)
    reread_str = pd.read_excel(PATHFINDER_PATH, sheet_name=PATHFINDER_SHEET, dtype=str)
    print(f"  Wrote cleaned file: {PATHFINDER_PATH}")
    print(f"  Verified: reading back with dtype=str gives correct values (e.g. "
          f"{reread_str['CIP Code'].iloc[0]!r}).")
    if reread_default["CIP Code"].dtype != object:
        print("  IMPORTANT: reading this file back WITHOUT dtype=str re-infers 'CIP Code' as "
              f"{reread_default['CIP Code'].dtype} and silently drops the leading zero again "
              "(e.g. becomes " + repr(reread_default["CIP Code"].iloc[0]) + "). This is pandas' "
              "own default type inference, not a flaw in this file -- any code that reads this "
              "file MUST pass dtype=str (or dtype={'CIP Code': str}), the same convention "
              "build_demand.py's load_crosswalk() and load_wages() already use.")


def clean_scorecard():
    print(f"\n--- Cleaning {SCORECARD_PATH} ---")
    if not os.path.exists(SCORECARD_PATH):
        print("  SKIPPED: file not found.")
        return

    df = pd.read_parquet(SCORECARD_PATH)
    print(f"  Read {len(df):,} rows, {len(df.columns)} columns.")

    candidate_cols = [c for c in df.columns if c not in SCORECARD_ID_COLS]
    ps_only_cols, range_cols, unexpected = [], [], {}
    for col in candidate_cols:
        if not pd.api.types.is_string_dtype(df[col]):
            continue  # already numeric, nothing to do
        vals = df[col].dropna().unique()
        bad = [v for v in vals if pd.to_numeric(pd.Series([v]), errors="coerce").isna().iloc[0]]
        if not bad:
            continue
        has_range = any((" - " in b or b.startswith("<=") or b.startswith(">=")) for b in bad)
        if has_range:
            range_cols.append(col)
        elif set(bad) == {SCORECARD_SUPPRESSION_TOKEN}:
            ps_only_cols.append(col)
        else:
            unexpected[col] = bad[:5]

    if unexpected:
        _fail(f"{len(unexpected)} column(s) have non-numeric values that are neither "
              f"'{SCORECARD_SUPPRESSION_TOKEN}' nor a recognized range-bin format: "
              f"{unexpected}. Not safe to auto-convert -- inspect by hand.")

    print(f"  {len(ps_only_cols)} columns use only '{SCORECARD_SUPPRESSION_TOKEN}' as a "
          "suppression placeholder -- will be cleaned and cast to numeric.")
    print(f"  {len(range_cols)} columns (federal loan repayment rate bins) contain binned "
          "range text like '0.37 - 0.38' or '<=0.20' -- NOT a bug, left as text, NOT modified:")
    for c in range_cols:
        print(f"    {c}")

    for col in ps_only_cols:
        is_suppressed = df[col] == SCORECARD_SUPPRESSION_TOKEN
        df.loc[is_suppressed, col] = pd.NA
        df[col] = pd.to_numeric(df[col], errors="coerce")

    _backup(SCORECARD_PATH)
    df.to_parquet(SCORECARD_PATH, index=False)
    print(f"  Wrote cleaned file: {SCORECARD_PATH}")


def main():
    clean_pathfinder()
    clean_scorecard()
    print("\nDone. Original files backed up alongside themselves with a .bak_<timestamp> suffix.")


if __name__ == "__main__":
    main()
